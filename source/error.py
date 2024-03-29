#Module that detects and autocorrects errors
import os
import re
import pygame
from openpyxl import load_workbook
from spellchecker import SpellChecker
from log import *

err = 0
warn = 0
fix = 0

#Function to search for and parse filename with regex
def parse_filename():
    global err, warn, fix
    dir = os.listdir(os.getcwd())
    #Check for valid excel file
    for file in dir:
        re_fn = re.search(r'(\w*)(\d\d\-\d\d\-\d+)(\.xlsx)', file)     #Filename is "<Alphanumeric and _>00-00-0*.xlsx"
        if re_fn != None:
            break
    if re_fn == None:
        err_filename((err+1, warn, fix))
    return re_fn

#Subroutine of detect() that checks sheetnames
def detect_sheetnames(workbook):
    global err, warn, fix
    sp = SpellChecker(language=None, distance=3, case_sensitive=True)
    true_names = ['Mice', 'Cages']
    for name in true_names:
        sp.word_frequency.add(name)
    for word in sp.unknown(workbook.sheetnames):
        new_name = sp.correction(word)
        if new_name is not word:                    #Make sure correction found
            workbook[word].title = new_name
            fix+=1
            st_autosheet(word, new_name)           #Log change to sheet name
    #Check if expected sheetnames all present
    for name in true_names:
        if name not in workbook.sheetnames:
            err_autosheet(name, (err+1, warn, fix))
    return

#Subroutine of detect() that checks column headers
def detect_headers(worksheet):
    global err, warn, fix
    if worksheet.title == 'Mice':
        true_names = ['Mouse ID', 'Cage ID', 'Ear Tag?', 'Sex', 'DOB', 'Age (days)', 'Pregnant?', 'Sacked Status: Potential (P), Sacked (S), Died (D)', 'Date of Death', 'Genotyped?', 'Runt?', 'Comments']
    elif worksheet.title == 'Cages':
        true_names = ['Cage ID', 'Status/Condition', 'Number of Pups', 'Pup DOB', 'Wean Date', 'Condition', 'Color']
    sp = SpellChecker(language=None, distance=3, case_sensitive=True)
    for name in true_names:
        sp.word_frequency.add(name)
    headers = worksheet['2']
    head_hash = {h.value:str(h.column)+str(h.row) for h in headers if h.value is not None}   #Generate hash of header:position
    for word in sp.unknown(head_hash.keys()):
        new_name = sp.correction(word)
        if new_name is not word:                    #Make sure correction found
            cell = worksheet[head_hash[word]]       #Select cell with incorrect word
            cell.value = new_name
            fix+=1
            st_autoheader(word, new_name)           #Log change to column header
    #Create list of current headers after autocorrection process
    curr_headers = [h.value for h in headers if h.value is not None]
    #Check if expected headers all present
    failed = False
    for name in true_names:
        if name not in curr_headers:
            failed = True
            err+=1
            err_autoheader(name)
    if failed:
        err_autocell_gen(worksheet.title, (err+1, warn, fix))
    return

#Mouse cell detection function; Iterates through selected columns and checks for blank cells
def detect_cells_m(worksheet):
    global err, warn, fix
    # Construct column headers dict (value:cell obj)
    headers = worksheet['2']
    col_dict = {h.value:h.column for h in headers if h.value is not None}
    failed = False     #bool to show when we have failed
    #Check Mouse IDs
    m_id_col = worksheet[col_dict['Mouse ID']][2:]      #Skip headers and blank cells
    for cell in m_id_col:
        if cell.value is None or str(cell.value).isspace():
            failed = True
            err+=1
            err_autocell(str(cell.column)+str(cell.row), cell.value, 'Mouse ID')       
    #Check Cage IDs
    cids = set([])
    c_id_col = worksheet[col_dict['Cage ID']][2:]
    for cell in c_id_col:
        if cell.value is None or str(cell.value).isspace():
            failed = True
            err+=1
            err_autocell(str(cell.column)+str(cell.row), cell.value, 'Cage ID')
        else:
            cids.add(cell.value)
    #Check Sex
    sex_col = worksheet[col_dict['Sex']][2:]
    for cell in sex_col:
        old_val = cell.value
        cell.value = str(cell.value).replace(' ', '')
        if str(cell.value).lower() not in ('f', 'm'):          
            cell.value = 'F'
            warn+=1
            warn_autocell(str(cell.column)+str(cell.row), old_val, cell.value, 'Sex')
    #Check Ages
    age_col = worksheet[col_dict['Age (days)']][2:]
    for cell in age_col:
        if cell.value is None or str(cell.value).isspace():  #Age missing or just whitespace
            old_val = cell.value
            cell.value = 0
            warn+=1
            warn_autocell(str(cell.column)+str(cell.row), old_val, cell.value, 'Age')
        elif not str(cell.value).isdigit():     #Non-digit chars in age (autocorrect)
            old_val = cell.value
            valid_chars = list(filter(str.isdigit, cell.value))
            if valid_chars:
                cell.value = "".join(valid_chars)
                fix+=1
                st_autocell(str(cell.column)+str(cell.row), old_val, cell.value, 'Age')
            else:
                cell.value = 0
                warn+=1
                warn_autocell(str(cell.column)+str(cell.row), old_val, cell.value, 'Age')
    #Check if sacked->date of death
    sac_col = worksheet[col_dict['Sacked Status: Potential (P), Sacked (S), Died (D)']][2:]
    for cell in sac_col:
        cell.value = str(cell.value).replace(' ', '')
        if str(cell.value).lower() in ('s', 'd'):   #if sac'd or died include date of death
            pos = str(col_dict['Date of Death']) + str(cell.row)
            dod_cell = worksheet[pos]
            if dod_cell.value is None or re.search(r'(\d\d\d\d)\-(\d\d)\-(\d\d)', str(dod_cell.value)) is None:
                old_val = dod_cell.value
                dod_cell.value = '0000-00-00'
                warn+=1
                warn_autocell(pos, old_val, dod_cell.value, 'Date of Death')
    if failed:
        err_autocell_gen('Mice', (err+1, warn, fix))
    return cids

#Cage cell detection function; Iterates through selected columns and checks for blank cells
def detect_cells_c(worksheet):
    global err, warn, fix
    # Construct column headers dict (value:cell obj)
    headers = worksheet['2']
    col_dict = {h.value:h.column for h in headers if h.value is not None}
    failed = False
    #Check Cage IDs
    cids = []
    c_id_col = worksheet[col_dict['Cage ID']][2:]
    for cell in c_id_col:
        if cell.value is None or str(cell.value).isspace():
            failed = True
            err+=1
            err_autocell(str(cell.column)+str(cell.row), cell.value, 'Cage ID')
        else:
            cids.append(cell.value)
    #Check if pups->DOB(required)->Wean Date(not required)
    pup_col = worksheet[col_dict['Number of Pups']][2:]
    for cell in pup_col:
        cell.value = str(cell.value).replace(' ', '')
        if cell.value is not None and str(cell.value).replace('.', '').isdigit() and int(cell.value) > 0:    # check if pups is a number greater than 0
            dob_pos = str(col_dict['Pup DOB']) + str(cell.row)
            dob_cell = worksheet[dob_pos]
            if dob_cell.value is None or re.search(r'(\d\d\d\d)\-(\d\d)\-(\d\d)', str(dob_cell.value)) is None:
                old_val = dob_cell.value
                dob_cell.value = '0000-00-00'
                warn+=1
                warn_autocell(dob_pos, old_val, dob_cell.value, 'Pup DOB')
            wd_pos = str(col_dict['Wean Date']) + str(cell.row)
            wd_cell = worksheet[wd_pos]
            if wd_cell.value and re.search(r'(\d\d\d\d)\-(\d\d)\-(\d\d)', str(wd_cell.value)) is None:
                old_val = wd_cell.value
                wd_cell.value = None
                warn+=1
                warn_autocell(wd_pos, old_val, wd_cell.value, 'Wean Date')
    #Check condition/color chart
    cond_col = worksheet[col_dict['Condition']][2:]
    cond_set = set([])
    color_list = pygame.color.THECOLORS.keys()
    for cond_cell in cond_col:
        if cond_cell.value is not None and not str(cond_cell.value).isspace():
            pos = str(col_dict['Color']) + str(cond_cell.row)
            color = worksheet[pos].value
            cond_set.add(str(cond_cell.value).lower())
            if color is None or str(color).isspace() or str(color).lower() not in color_list:
                failed = True
                err+=1
                err_cond_color(pos, color)
    #Check status/condition column for unrecognized condition
    stat_col = worksheet[col_dict['Status/Condition']][2:]
    for stat_cell in stat_col:
        if stat_cell.value is not None and not str(stat_cell.value).isspace():
            if str(stat_cell.value).lower() not in cond_set:
                pos = str(stat_cell.column)+str(stat_cell.row)
                old_val = stat_cell.value
                stat_cell.value = None
                warn+=1
                warn_autocell(pos, old_val, stat_cell.value, 'Status/Condition')
    if failed:
        err_autocell_gen('Cages', (err+1, warn, fix))
    return cids

#Function to simply delete the blank rows from the excel sheet to avoid confusion with blank cells
def delete_blank_rows(worksheet):
    for row in worksheet.iter_rows():
        if not any(cell.value for cell in row): #Row is completely blank
            worksheet.delete_rows(row[0].row, 1)
            delete_blank_rows(worksheet)
            return

#Function to compare list of cages generated by mouse sheet and cage sheet
def compare_cage_lists(cids_m, cids_c):
    global err, warn, fix
    failed = False
    set_cids_m, set_cids_c = cids_m, set(cids_c)    #cids_m is already a set
    if len(set_cids_c) < len(cids_c):   #Check for duplicate CIDS in cage sheet
        for cid in set_cids_c:
            if cids_c.count(cid) > 1:
                err+=1
                err_dup_cid(cid)
        failed = True
    #Error: Mice diff Cages cids
    diff = set_cids_m.difference(set_cids_c)
    if diff:
        for cid in diff:
            err+=1
            err_miss_cage(cid)
        failed = True          
    #Warning: Cages diff Mice cids
    diff = set_cids_c.difference(set_cids_m)
    if diff:
        for cid in diff:
            warn+=1
            warn_empty_cage(cid)
    if failed:
        err_autocell_gen('Cages', (err+1, warn, fix))
    return

#Main error detection function: Calls subroutines for checking and correcting/logging sheet names, column headers, and cell values
def detect(filename):
    global err, warn, fix
    wb = load_workbook(filename)        #Load workbook
    st_detect()
    detect_sheetnames(wb)          
    ws_m, ws_c = wb['Mice'], wb['Cages']   #Load worksheets
    ws_m_max, ws_c_max = ws_m.max_row, ws_c.max_row
    delete_blank_rows(ws_m)
    delete_blank_rows(ws_c)
    if(ws_m_max - ws_m.max_row > 0):
        st_del_rows(ws_m_max-ws_m.max_row, 'Mice')
        fix+=1
    if(ws_c_max - ws_c.max_row > 0):
        st_del_rows(ws_c_max - ws_c.max_row, 'Cages')
        fix+=1
    detect_headers(ws_m)
    detect_headers(ws_c)
    cids_m = detect_cells_m(ws_m)
    cids_c = detect_cells_c(ws_c)
    compare_cage_lists(cids_m, cids_c)      #Detect mismatch/duplicate CIDs between Mice vs. Cages sheet
    wb.save('temp.xlsx')                 #Save temp file
    st_ewf((err, warn, fix))
    return
